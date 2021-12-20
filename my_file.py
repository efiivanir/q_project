import qinvest as qi
from qinvest.Common import Common as co
import pandas as _pd
import numpy as _np
import yfinance as _yf
from yahoo_fin import stock_info as _si
from openpyxl import Workbook, worksheet, load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side

#
# sp = qi.Finance.get_brk_portfolio()

# co.cleanExcel("SPY.xlsx")

work_book = load_workbook(filename='RDWR.xlsx', read_only=False)
sheet_names = work_book.sheetnames
sheet = work_book['actions']
colummn = sheet['A']
for cell in colummn:
    cord = cell.coordinate
    s = sheet[cord].value
    if s == 'Date':
        continue
    sheet[cord].value = s.strftime("%Y-%m-%d")
