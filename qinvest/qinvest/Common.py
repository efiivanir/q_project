from openpyxl import Workbook, worksheet,load_workbook
from openpyxl.styles import Font, Color, Alignment, Border, Side
import datetime
class Common():
    def whitespace_remover(dataframe):
        # iterating over the columns
        for i in dataframe.columns:

            # checking datatype of each columns
            if dataframe[i].dtype == 'object':

                # applying strip function on column
                dataframe[i] = dataframe[i].map(str.strip)
            else:

                # if condn. is False then it will do nothing.
                pass
        return dataframe.copy()

    def cleanExcel(file_path):
        work_book = load_workbook(filename=file_path, read_only=False)
        sheet_names = work_book.sheetnames
        for sh in sheet_names:
            sheet = work_book[sh]
            sheet.freeze_panes = 'B2'
            left_aligned_text = Alignment(horizontal="left")
            fontStyle = Font(size="16")
            row_count = sheet.max_row
            column_count = sheet.max_column
            for i in range(1, row_count + 1):
                for j in range(1, column_count + 1):
                    c = sheet.cell(row=i, column=j)
                    c.alignment = left_aligned_text
                    c.font = fontStyle
                    val = c.value
                    if isinstance(val, datetime.date):
                        c.value = val.strftime("%Y-%m-%d")

            if sh == 'history':
                columns = ('B','C','D','E','I')
                for col in columns:
                    colummn = sheet[col]
                    for cell in colummn:
                        cord = cell.coordinate
                        if cord == col + '1':
                            continue
                        sheet[cord].number_format = '0.000'

            dims = {}
            for row in sheet.rows:
                for cell in row:
                    if cell.value:
                        dims[cell.column_letter] = max((dims.get(cell.column_letter, 0),
                                                        len(str(cell.value))))
            for col, value in dims.items():
                sheet.column_dimensions[col].width = value + 1

        work_book.save(file_path)
        work_book.close()
